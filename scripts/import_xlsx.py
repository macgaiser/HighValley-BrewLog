"""Importiert historische Sude und den Lagerbestand aus der ursprünglichen
Excel-Vorlage in die Datenbank.

Es werden ausschließlich Werte importiert, keine Formeln - alle Kennzahlen
(IBU, Alkohol, Ausbeute, ...) werden von der App selbst neu berechnet
(app/formulas.py, app/batch_calc.py), nicht aus der Excel übernommen, da
mehrere Formeln dort nachweislich fehlerhaft sind (u.a. eine #NAME?-Formel
für den Alkoholgehalt, die in jedem Sud auftritt, und eine vertauschte Zelle
in "Batch #57 Festbier", die dort die Stammwürze-Berechnung zerstört).

Aufruf:
    python scripts/import_xlsx.py /pfad/zur/Brauprotokoll.xlsx

Der Import ist idempotent bezüglich des Blattnamens: ein bereits
importiertes Blatt (erkannt an einem Kommentar "Importiert aus Excel-Tab
'<Blattname>'") wird beim erneuten Aufruf übersprungen - das Skript kann
also gefahrlos mehrfach und mit mehreren Dateiversionen aufgerufen werden
(z.B. einmal für das aktuelle Braubuch, einmal für ein älteres Archiv mit
weiteren historischen Suden).

Das Braubuch-Layout hat sich über die Jahre verändert (frühe Sude sind
deutlich knapper dokumentiert als spätere, u.a. ohne Alphasäure-Angabe bei
Hopfengaben und ohne strukturierten Gärverlauf). Das Skript erkennt das
jeweilige Layout je Tabellenblatt automatisch und importiert, was jeweils
vorhanden ist; für ältere Sude bleiben manche berechnete Kennzahlen (IBU,
Alkohol, Ausbeute) daher leer, weil die dafür nötigen Rohdaten im
Original fehlen - die damals von Hand eingetragenen Werte werden dafür als
Kommentar "Original-Aufzeichnung im Braubuch: ..." übernommen.
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sqlmodel import Session, select

from app.database import engine, init_db
from app.models import (
    Batch,
    BatchComment,
    BrewDayTask,
    CarbonationEntry,
    DryHopAddition,
    FermentationLogEntry,
    GrainAddition,
    HopAddition,
    HopAdditionType,
    InventoryCategory,
    InventoryItem,
    MashStep,
    YeastAddition,
)

IMPORT_MARKER = "Importiert aus Excel-Tab"

# Bekannte Tippfehler/Datenfehler in der Original-Excel, die sich nicht aus
# der Datei selbst korrigieren lassen (z.B. weil dort versehentlich ein
# Datum statt einer Messung steht). Key = exakter Blattname, Value = vom
# Nutzer bestätigte Korrektur.
KNOWN_CORRECTIONS: dict[str, dict[str, float]] = {
    "Batch #57 Festbier": {
        # In der Excel steht in der Zelle für "Stammwürze nach Kochen und
        # Kühlung" ein Datum statt eines Brix-Werts. Vom Nutzer bestätigt (29.08.2026): 13,6 °Brix.
        "post_boil_brix": 13.6,
    },
}


def col_label_rows(ws, column: str, max_row: int = 120) -> dict[str, int]:
    col_idx = openpyxl.utils.column_index_from_string(column)
    labels: dict[str, int] = {}
    for r in range(1, max_row + 1):
        value = ws.cell(row=r, column=col_idx).value
        if isinstance(value, str) and value.strip():
            labels.setdefault(value.strip(), r)
    return labels


EXCEL_EPOCH = datetime(1899, 12, 30)


def as_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime) and value.year in (1899, 1900):
        # In ein paar Zellen wurde eine reine Zahl (z.B. "24,5") eingetippt,
        # die Excel wegen der Zellformatierung stattdessen als Datum nahe
        # seiner Epoche interpretiert hat. Über die Seriennummer lässt sich
        # der ursprünglich gemeinte Zahlenwert zurückrechnen.
        delta = value - EXCEL_EPOCH
        return delta.days + delta.seconds / 86400
    if isinstance(value, str):
        match = re.search(r"-?\d+[.,]?\d*", value)
        if match:
            return float(match.group(0).replace(",", "."))
    return None


def as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def as_sum(value) -> float | None:
    """Wie as_float, summiert aber mehrere Zahlen in einem Feld (z.B. die
    ganz frühen Sude tragen die Ausschlagwürze manchmal als "15+2,5" ein,
    Hauptmenge + separat nachgeschüttetes Wasser)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        numbers = re.findall(r"\d+[.,]?\d*", value)
        if numbers:
            return sum(float(n.replace(",", ".")) for n in numbers)
    return None


def cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def find_label(a_labels: dict[str, int], *aliases: str) -> int | None:
    """Das Braubuch-Layout hat sich über die Jahre leicht verändert (z.B.
    "Wasserausgleich" vs. "Wasserausgleich:", "Brauvorgang Stonstiges:" vs.
    nur "Stonstiges:" in älteren Suden) - hier werden bekannte
    Schreibweisen-Varianten eines Abschnitts durchprobiert."""
    for alias in aliases:
        if alias in a_labels:
            return a_labels[alias]
    return None


def import_batch_sheet(ws, session: Session) -> Batch | None:
    a_labels = col_label_rows(ws, "A")
    l_labels = col_label_rows(ws, "L")

    match = re.search(r"#(\d+)", ws.title)
    if not match:
        return None
    batch_number = int(match.group(1))

    # Ältere Sude (grob #1-#49 im ersten Braubuch) haben im Würzekochen-Block
    # keine Alphasäure-Spalte - dort wurde IBU nur als geschätzte Summe in
    # "Bittere:" eingetragen statt je Hopfengabe berechnet. Erkennung anhand
    # der Kopfzeile direkt in der "Würzekochen"-Zeile (Spalte C: "Alpha"/
    # "Anpha" bei neuerem Format, "Menge" beim alten).
    hop_header_row = a_labels.get("Würzekochen")
    hop_header_text = str(cell(ws, hop_header_row, 3) or "") if hop_header_row else ""
    legacy_hop_layout = "lpha" not in hop_header_text

    batch = Batch(batch_number=batch_number)
    batch.name = cell(ws, 2, 2) or ws.title
    batch.brew_date = as_date(cell(ws, 1, 5))
    batch.fermentation_type = (cell(ws, 3, 5) or "").strip()
    batch.target_volume_l = as_sum(cell(ws, 4, 2))
    batch.color_ebc = as_float(cell(ws, 8, 2))

    if "Hauptguss" in a_labels:
        r = a_labels["Hauptguss"]
        batch.main_water_l = as_float(cell(ws, r, 5))
    if "Nachguss" in a_labels:
        r = a_labels["Nachguss"]
        batch.sparge_water_l = as_float(cell(ws, r, 5))
    if "Milchsäure 80%" in a_labels:
        r = a_labels["Milchsäure 80%"]
        batch.lactic_acid_80_ml = as_float(cell(ws, r, 5))

    # Kochzeit steht im alten Layout in Spalte E, im neuen in Spalte F (weil
    # dort zusätzlich eine Alphasäure-Spalte eingeschoben wurde).
    if "Kochzeit" in a_labels:
        boil_col = 5 if legacy_hop_layout else 6
        batch.boil_time_min = as_float(cell(ws, a_labels["Kochzeit"], boil_col))

    if "Stammwürze vor Läutern:" in a_labels:
        batch.pre_lauter_brix = as_float(cell(ws, a_labels["Stammwürze vor Läutern:"], 5))
    if "Stammwürze nach Läutern:" in a_labels:
        r = a_labels["Stammwürze nach Läutern:"]
        batch.post_lauter_brix = as_float(cell(ws, r, 5))
        batch.post_lauter_volume_l = as_float(cell(ws, r, 7))
    wasserausgleich_row = find_label(a_labels, "Wasserausgleich", "Wasserausgleich:")
    if wasserausgleich_row:
        batch.water_adjustment_l = as_float(cell(ws, wasserausgleich_row, 5))
    if "Stammwürze nach Kochen und Kühlung" in a_labels:
        r = a_labels["Stammwürze nach Kochen und Kühlung"]
        raw = cell(ws, r, 5)
        # Batch #57 Festbier: hier steht in der Original-Excel versehentlich
        # ein Datum statt der Brix-Messung -> nicht importieren.
        if isinstance(raw, (int, float)):
            batch.post_boil_brix = float(raw)
        batch.post_boil_volume_l = as_float(cell(ws, r, 7))

    for field, value in KNOWN_CORRECTIONS.get(ws.title, {}).items():
        setattr(batch, field, value)

    # Neueres Layout: geplante Stammwürze in E6 ("geplant:" in D6). Älteres
    # Layout kennt kein separates "gemessen nach Kochen" -> B6 ("Stammwürze:")
    # ist dort der einzige verfügbare Wert und dient als Fallback.
    batch.target_og_plato = as_float(cell(ws, 6, 5)) or as_float(cell(ws, 6, 2))

    session.add(batch)
    session.commit()
    session.refresh(batch)

    # Schüttung (ganz frühe Sude haben keine "Summe:"-Zeile -> Ende der
    # Maischplan-Abschnitt als Fallback-Grenze)
    grain_end = find_label(a_labels, "Summe:") or a_labels.get("Maischplan")
    if "Schüttung:" in a_labels and grain_end:
        start, end = a_labels["Schüttung:"] + 1, grain_end
        pos = 0
        for r in range(start, end):
            name = cell(ws, r, 1)
            amount = as_float(cell(ws, r, 5))
            if name and amount:
                session.add(GrainAddition(batch_id=batch.id, position=pos, malt_name=str(name).strip(), amount_kg=amount))
                pos += 1

    # Maischplan
    if "Maischplan" in a_labels and "Würzekochen" in a_labels:
        start, end = a_labels["Maischplan"] + 1, a_labels["Würzekochen"]
        pos = 0
        for r in range(start, end):
            name = cell(ws, r, 1)
            if name:
                session.add(
                    MashStep(
                        batch_id=batch.id,
                        position=pos,
                        name=str(name).strip(),
                        temperature_c=as_float(cell(ws, r, 3)),
                        duration_min=as_float(cell(ws, r, 5)),
                        comment=str(cell(ws, r, 7) or "").strip(),
                    )
                )
                pos += 1

    # Würzekochen / Hopfengaben. "Hefegabe" folgt in beiden Layout-Versionen
    # direkt auf den Hopfenblock und ist eine zuverlässigere Endgrenze als
    # der "Brauvorgang Stonstiges:"-Abschnitt, der nicht in jedem Sud existiert.
    if "Kochzeit" in a_labels and "Hefegabe" in a_labels:
        start, end = a_labels["Kochzeit"] + 1, a_labels["Hefegabe"]
        pos = 0
        if legacy_hop_layout:
            amount_col, alpha_col, time_col, temp_col = 3, None, 5, 7
        else:
            amount_col, alpha_col, time_col, temp_col = 4, 3, 6, 8
        for r in range(start, end):
            step_label = str(cell(ws, r, 1) or "")
            amount = as_float(cell(ws, r, amount_col))
            hop_name = cell(ws, r, 2)
            if amount is None or not step_label:
                continue
            addition_type = HopAdditionType.kochen
            if "whirlpool" in step_label.lower():
                addition_type = HopAdditionType.whirlpool
            elif "nachisomerisierung" in step_label.lower():
                addition_type = HopAdditionType.nachisomerisierung
            alpha = as_float(cell(ws, r, alpha_col)) if alpha_col else None
            session.add(
                HopAddition(
                    batch_id=batch.id,
                    position=pos,
                    hop_name=str(hop_name).strip() if hop_name else step_label,
                    alpha_acid_percent=alpha * 100 if alpha is not None else None,
                    amount_g=amount,
                    time_min=as_float(cell(ws, r, time_col)),
                    temperature_c=as_float(cell(ws, r, temp_col)),
                    addition_type=addition_type,
                )
            )
            pos += 1

    # Hefegabe
    if "Hefegabe" in a_labels and "Stopfhopfen" in a_labels:
        start, end = a_labels["Hefegabe"] + 1, a_labels["Stopfhopfen"]
        pos = 0
        for r in range(start, end):
            name = cell(ws, r, 1)
            if name:
                session.add(
                    YeastAddition(
                        batch_id=batch.id,
                        position=pos,
                        yeast_name=str(name).strip(),
                        generation_label=str(cell(ws, r, 2) or "").strip(),
                        amount=as_float(cell(ws, r, 3)),
                        unit=str(cell(ws, r, 4) or "g").strip() or "g",
                        pitch_temperature_c=as_float(cell(ws, r, 7)),
                    )
                )
                pos += 1

    # Stopfhopfen
    if "Stopfhopfen" in a_labels and "Karbonisierung" in a_labels:
        start, end = a_labels["Stopfhopfen"] + 1, a_labels["Karbonisierung"]
        pos = 0
        for r in range(start, end):
            name = cell(ws, r, 1)
            amount = as_float(cell(ws, r, 3))
            if name and amount:
                session.add(
                    DryHopAddition(
                        batch_id=batch.id,
                        position=pos,
                        hop_name=str(name).strip(),
                        timing_label=str(cell(ws, r, 2) or "").strip(),
                        amount_g=amount,
                    )
                )
                pos += 1

    # Karbonisierung. Je nach Braubuch-Version folgt danach der Gärverlauf,
    # ein "(Brauvorgang) Stonstiges:"-Abschnitt oder direkt die Kommentare.
    carb_end = min(
        (
            r
            for r in (
                a_labels.get("Gärung - Datum"),
                find_label(a_labels, "Brauvorgang Stonstiges:", "Stonstiges:"),
                a_labels.get("Kommentar:"),
            )
            if r and a_labels.get("Karbonisierung") and r > a_labels["Karbonisierung"]
        ),
        default=None,
    )
    if "Karbonisierung" in a_labels and carb_end:
        start, end = a_labels["Karbonisierung"] + 1, carb_end
        pos = 0
        for r in range(start, end):
            sugar = as_float(cell(ws, r, 3))
            if sugar:
                session.add(
                    CarbonationEntry(
                        batch_id=batch.id,
                        position=pos,
                        sugar_g=sugar,
                        bottle_volume_l=as_float(cell(ws, r, 5)) or 0.5,
                        label=str(cell(ws, r, 1) or "Zucker").strip(),
                    )
                )
                pos += 1

    # Gärverlauf
    if "Gärung - Datum" in a_labels:
        start = a_labels["Gärung - Datum"] + 1
        end = a_labels.get("Kommentar:", start + 15)
        pos = 0
        for r in range(start, end):
            brix = as_float(cell(ws, r, 2))
            if brix is None:
                continue
            session.add(
                FermentationLogEntry(
                    batch_id=batch.id,
                    position=pos,
                    entry_date=as_date(cell(ws, r, 1)),
                    brix=brix,
                    comment=str(cell(ws, r, 6) or "").strip(),
                )
            )
            pos += 1

    # Brautag-Zeitplan (Spalte L)
    if "Tätigkeit" in l_labels:
        start = l_labels["Tätigkeit"] + 1
        end = l_labels.get("Summe:", start + 20)
        pos = 0
        for r in range(start, end):
            name = cell(ws, r, 12)
            duration = as_float(cell(ws, r, 15))
            if name:
                session.add(
                    BrewDayTask(
                        batch_id=batch.id,
                        position=pos,
                        task_name=str(name).strip(),
                        planned_duration_min=duration,
                        note=str(cell(ws, r, 17) or "").strip(),
                    )
                )
                pos += 1

    # Kommentare
    pos = 0
    if "Kommentar:" in a_labels:
        for r in range(a_labels["Kommentar:"] + 1, ws.max_row + 1):
            text = cell(ws, r, 1)
            if text:
                session.add(BatchComment(batch_id=batch.id, position=pos, text=str(text).strip()))
                pos += 1

    if legacy_hop_layout:
        eff, ibu, abv = cell(ws, 5, 2), cell(ws, 7, 2), cell(ws, 9, 2)
        parts = []
        if isinstance(eff, (int, float)):
            parts.append(f"Ausbeute ~{eff:g}%")
        if isinstance(ibu, (int, float)):
            parts.append(f"Bittere ~{ibu:g} IBU")
        if abv is not None:
            parts.append(f"Alkohol ~{abv}" + (" Vol.-%" if isinstance(abv, (int, float)) else ""))
        if parts:
            session.add(
                BatchComment(
                    batch_id=batch.id,
                    position=pos,
                    text=(
                        "Original-Aufzeichnung im Braubuch: "
                        + ", ".join(parts)
                        + " (historischer Wert, nicht neu berechnet - Alphasäure der "
                        "Hopfengaben bzw. Gärverlaufsmessungen fehlen im Originalprotokoll "
                        "für diesen Sud)."
                    ),
                )
            )
            pos += 1

    if ws.title in KNOWN_CORRECTIONS:
        session.add(
            BatchComment(
                batch_id=batch.id,
                position=pos,
                text=(
                    "Import-Hinweis: Stammwürze nach Kochen/Kühlung war in der "
                    "Original-Excel fehlerhaft (Datum statt Messwert) und wurde "
                    "auf 13,6 °Brix korrigiert."
                ),
            )
        )
        pos += 1

    session.add(
        BatchComment(
            batch_id=batch.id,
            position=pos,
            text=f"{IMPORT_MARKER} '{ws.title}'",
        )
    )

    session.commit()
    return batch


def already_imported(session: Session, sheet_title: str) -> bool:
    stmt = select(BatchComment).where(BatchComment.text == f"{IMPORT_MARKER} '{sheet_title}'")
    return session.exec(stmt).first() is not None


def import_inventory_sheet(ws, session: Session) -> int:
    """Importiert den Lagerbestand-Tab. Wird nur ausgeführt, wenn noch keine
    Lagerartikel vorhanden sind (idempotent für den Erstimport)."""

    if session.exec(select(InventoryItem)).first() is not None:
        return 0

    count = 0

    def add_item(category: InventoryCategory, name, brand, spec, amount, unit):
        nonlocal count
        if not name:
            return
        item = InventoryItem(
            category=category,
            name=str(name).strip(),
            brand=str(brand or "").strip(),
            spec=str(spec or "").strip(),
            amount=as_float(amount) or 0,
            unit=str(unit or "kg").strip() or "kg",
        )
        session.add(item)
        count += 1

    a_labels = col_label_rows(ws, "A", max_row=200)

    # Malz: Zeilen 4..25 (bis "Hopfen:")
    if "Malz:" in a_labels and "Hopfen:" in a_labels:
        for r in range(a_labels["Malz:"] + 1, a_labels["Hopfen:"]):
            name = cell(ws, r, 1)
            if name:
                add_item(InventoryCategory.malz, name, cell(ws, r, 2), cell(ws, r, 3), cell(ws, r, 5), cell(ws, r, 6))

    if "Hopfen:" in a_labels and "Hefe" in a_labels:
        for r in range(a_labels["Hopfen:"] + 1, a_labels["Hefe"]):
            name = cell(ws, r, 1)
            if name:
                add_item(InventoryCategory.hopfen, name, cell(ws, r, 2), cell(ws, r, 3), cell(ws, r, 5), cell(ws, r, 6))

    if "Hefe" in a_labels:
        for r in range(a_labels["Hefe"] + 1, ws.max_row + 1):
            name = cell(ws, r, 1)
            if name:
                add_item(InventoryCategory.hefe, name, cell(ws, r, 2), cell(ws, r, 3), cell(ws, r, 5), cell(ws, r, 6))

    session.commit()
    return count


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--skip-inventory", action="store_true")
    args = parser.parse_args()

    init_db()
    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    imported, skipped = 0, 0
    with Session(engine) as session:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name.strip().lower() == "lagerbestand":
                if not args.skip_inventory:
                    n = import_inventory_sheet(ws, session)
                    print(f"Lagerbestand: {n} Artikel importiert (0 bedeutet: bereits vorhanden, übersprungen)")
                continue
            if not re.search(r"#\d+", sheet_name):
                # z.B. "#31 Cider" statt "Batch #31 Cider" - Titel ohne
                # "Batch"-Präfix, aber mit Sud-Nummer, kommt im älteren
                # Braubuch vor.
                continue
            if already_imported(session, sheet_name):
                print(f"übersprungen (bereits importiert): {sheet_name}")
                skipped += 1
                continue
            batch = import_batch_sheet(ws, session)
            if batch:
                print(f"importiert: {sheet_name} -> Sud #{batch.batch_number} (id={batch.id})")
                imported += 1

    print(f"\nFertig. {imported} Sude importiert, {skipped} übersprungen.")


if __name__ == "__main__":
    main()
